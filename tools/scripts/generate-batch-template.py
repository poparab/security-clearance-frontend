from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import csv

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))

wb = Workbook()
ws = wb.active
ws.title = "Batch Template"
ws.sheet_view.rightToLeft = False  # LTR for data entry

# ── Column definitions ──────────────────────────────────────────────────────
COLUMNS = [
    ("Traveler_Name",           "Required",    "Full name as in passport (2-150 chars)"),
    ("Nationality",             "Required",    "Nationality — must match configured list"),
    ("Passport_Number",         "Required",    "6-20 alphanumeric characters"),
    ("Passport_Expiry",         "Required",    "YYYY-MM-DD — must be 6+ months from Travel_Date"),
    ("Travel_Date",             "Required",    "YYYY-MM-DD — today or future"),
    ("Departure_Country",       "Required",    "Country of departure"),
    ("Arrival_Airport_Egypt",   "Required",    "Select from configured Egyptian airport list"),
    ("Purpose",                 "Required",    "Tourism / Business / Medical / Family / Education / Conference / Transit / Other"),
    ("Flight_Number",           "Optional",    "Max 10 alphanumeric chars"),
    ("Transit_Stop_1_Country",  "Conditional", "Required if stop 1 exists"),
    ("Transit_Stop_1_Airport",  "Optional",    "Airport code or name for stop 1"),
    ("Transit_Stop_1_Hours",    "Optional",    "Layover duration 0-72 hours"),
    ("Transit_Stop_1_Flight",   "Optional",    "Connecting flight number for stop 1"),
    ("Transit_Stop_2_Country",  "Conditional", "Required if stop 2 exists"),
    ("Transit_Stop_2_Airport",  "Optional",    "Airport code or name for stop 2"),
    ("Transit_Stop_2_Hours",    "Optional",    "Layover duration 0-72 hours"),
    ("Transit_Stop_2_Flight",   "Optional",    "Connecting flight number for stop 2"),
    ("Transit_Stop_3_Country",  "Conditional", "Required if stop 3 exists"),
    ("Transit_Stop_3_Airport",  "Optional",    "Airport code or name for stop 3"),
    ("Transit_Stop_3_Hours",    "Optional",    "Layover duration 0-72 hours"),
    ("Transit_Stop_3_Flight",   "Optional",    "Connecting flight number for stop 3"),
]

# ── Styles ──────────────────────────────────────────────────────────────────
header_fill    = PatternFill("solid", fgColor="1D4ED8")   # primary blue
guidance_fill  = PatternFill("solid", fgColor="EFF6FF")   # primary light
req_fill       = PatternFill("solid", fgColor="FEE2E2")   # danger light
cond_fill      = PatternFill("solid", fgColor="FEF3C7")   # warning light
opt_fill       = PatternFill("solid", fgColor="F8FAFC")   # bg-alt

header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
guidance_font  = Font(name="Calibri", italic=True, color="1E40AF", size=10)
data_font      = Font(name="Calibri", size=11)

center_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin_side      = Side(style="thin", color="CBD5E1")
thin_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# ── Row 1: Headers ───────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 28
for col_idx, (col_name, req, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font      = header_font
    cell.fill      = header_fill
    cell.alignment = center_align
    cell.border    = thin_border

# ── Row 2: Guidance notes ────────────────────────────────────────────────────
ws.row_dimensions[2].height = 48
for col_idx, (_, req, guidance) in enumerate(COLUMNS, start=1):
    fill = req_fill if req == "Required" else (cond_fill if req == "Conditional" else opt_fill)
    cell = ws.cell(row=2, column=col_idx, value=f"[{req}] {guidance}")
    cell.font      = guidance_font
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = thin_border

# ── Rows 3–52: Data entry rows (50 rows) ─────────────────────────────────────
for row in range(3, 53):
    ws.row_dimensions[row].height = 20
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font      = data_font
        cell.alignment = left_align
        cell.border    = thin_border

# ── Column widths ────────────────────────────────────────────────────────────
COL_WIDTHS = [
    22, 18, 20, 18, 14, 22, 26, 14, 14,
    22, 22, 14, 18,
    22, 22, 14, 18,
    22, 22, 14, 18,
]
for i, width in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ── Freeze top two rows ──────────────────────────────────────────────────────
ws.freeze_panes = "A3"

# ── Sheet tab colour ─────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = "1D4ED8"

# ── Save xlsx ────────────────────────────────────────────────────────────────
output_dir = os.path.join(PROJECT_ROOT, "app", "assets")
os.makedirs(output_dir, exist_ok=True)

output_path = os.path.join(output_dir, "batch-template.xlsx")
wb.save(output_path)
print(f"Template saved to: {output_path}")
print(f"  Columns: {len(COLUMNS)}")
print(f"  Data rows: 50 (rows 3-52)")

# ── Save CSV ─────────────────────────────────────────────────────────────────
csv_path = os.path.join(output_dir, "batch-template.csv")
with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow([col[0] for col in COLUMNS])
    writer.writerow([f"[{col[1]}] {col[2]}" for col in COLUMNS])
print(f"CSV saved to: {csv_path}")
